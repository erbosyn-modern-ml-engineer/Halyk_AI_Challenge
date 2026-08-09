"""I/O helpers for Stage 5B routing outputs."""

from __future__ import annotations

import csv
import io
import json
import re
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from halyk_agent.domain.evidence import EvidenceSpan
from halyk_agent.domain.routing.models import (
    ConflictKind,
    DiagnosticSeverity,
    LedgerRow,
    RoutingReport,
)

_FIELD_ALIASES = {
    "txn_id": {"txn_id", "transaction_id", "transaction id", "txn id", "transactionid"},
    "date": {"date", "transaction_date", "transaction date", "txn_date", "posting_date"},
    "account_id": {"account_id", "account id", "account", "account_code", "account code"},
    "counterparty": {"counterparty", "counter_party", "vendor", "customer", "контрагент"},
    "description": {"description", "details", "narrative", "memo", "описание", "назначение"},
    "amount": {"amount", "value", "transaction_amount", "transaction amount", "сумма"},
    "currency": {"currency", "ccy", "currency_code", "currency code", "валюта"},
}
_REQUIRED = {"txn_id", "date", "account_id", "counterparty", "amount", "currency"}


class RoutingIOError(Exception):
    def __init__(self, message: str, *, code: str = "ROUTING_IO") -> None:
        super().__init__(message)
        self.message = message
        self.code = code


def _atomic_write(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8", newline="\n")
    tmp.replace(path)


def _header_key(value: object) -> str:
    text = str(value or "").strip().casefold()
    return re.sub(r"[\s\-]+", "_", text).strip("_")


def _header_mapping(header: list[object]) -> dict[str, int]:
    normalized = [_header_key(value) for value in header]
    mapping: dict[str, int] = {}
    for canonical, aliases in _FIELD_ALIASES.items():
        accepted = {_header_key(alias) for alias in aliases}
        matches = [index for index, name in enumerate(normalized) if name in accepted]
        if len(matches) > 1:
            raise RoutingIOError(
                f"ledger has multiple columns for {canonical}: {matches}",
                code="LEDGER_SCHEMA",
            )
        if matches:
            mapping[canonical] = matches[0]
    if not _REQUIRED.issubset(mapping):
        missing = sorted(_REQUIRED - set(mapping))
        raise RoutingIOError(f"ledger missing required columns: {missing}", code="LEDGER_SCHEMA")
    return mapping


def _row_value(row: list[object], mapping: dict[str, int], name: str) -> str:
    index = mapping.get(name)
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _rows_from_matrix(
    rows: Iterable[list[object]],
    *,
    source_file: str,
) -> tuple[LedgerRow, ...]:
    iterator = iter(rows)
    header = next(iterator, None)
    if header is None:
        raise RoutingIOError("ledger has no header row", code="LEDGER_SCHEMA")
    mapping = _header_mapping(header)
    out: list[LedgerRow] = []
    for index, row in enumerate(iterator):
        if not any(str(value or "").strip() for value in row):
            continue
        out.append(
            LedgerRow(
                row_index=index,
                txn_id=_row_value(row, mapping, "txn_id"),
                date=_row_value(row, mapping, "date"),
                account_id=_row_value(row, mapping, "account_id"),
                counterparty=_row_value(row, mapping, "counterparty"),
                description=_row_value(row, mapping, "description"),
                amount=_row_value(row, mapping, "amount"),
                currency=_row_value(row, mapping, "currency"),
                ledger_source_file=source_file,
            )
        )
    return tuple(out)


def load_ledger_csv(path: Path) -> tuple[LedgerRow, ...]:
    """Load a supported primary ledger into typed rows."""
    if not path.is_file():
        raise RoutingIOError(f"ledger not found: {path}", code="MISSING_LEDGER")
    data = path.read_bytes()
    suffix = path.suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        return load_ledger_xlsx_bytes(data, source_file=path.name)
    return load_ledger_csv_bytes(data, source_file=path.name)


def load_ledger_csv_bytes(data: bytes, *, source_file: str) -> tuple[LedgerRow, ...]:
    """Parse delimited ledger bytes with BOM and common header aliases."""
    stable_source_file = source_file.replace("\\", "/").rsplit("/", 1)[-1]
    if not stable_source_file:
        raise RoutingIOError("ledger source_file has no basename", code="LEDGER_SOURCE")
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise RoutingIOError(f"ledger is not UTF-8: {exc}", code="LEDGER_ENCODING") from exc
    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    matrix = [list(row) for row in csv.reader(io.StringIO(text), dialect=dialect)]
    return _rows_from_matrix(matrix, source_file=stable_source_file)


def load_ledger_xlsx_bytes(data: bytes, *, source_file: str) -> tuple[LedgerRow, ...]:
    """Parse the active worksheet of an XLSX ledger without formula evaluation."""
    stable_source_file = source_file.replace("\\", "/").rsplit("/", 1)[-1]
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise RoutingIOError(f"invalid XLSX ledger: {exc}", code="LEDGER_SCHEMA") from exc
    try:
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    return _rows_from_matrix(rows, source_file=stable_source_file)


def load_template_answers(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return _template_answers_from_payload(payload)


def load_template_answers_bytes(data: bytes) -> dict[str, Any]:
    try:
        payload = json.loads(data.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise RoutingIOError(
            f"invalid submission template JSON: {exc}", code="TEMPLATE_SCHEMA"
        ) from exc
    return _template_answers_from_payload(payload)


def _template_answers_from_payload(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict) or "answers" not in payload:
        raise RoutingIOError("submission template missing answers", code="TEMPLATE_SCHEMA")
    answers = payload["answers"]
    if not isinstance(answers, dict):
        raise RoutingIOError(
            "submission template answers must be an object",
            code="TEMPLATE_SCHEMA",
        )
    return answers


def load_evidence_catalogue(path: Path) -> tuple[EvidenceSpan, ...]:
    if not path.is_file():
        return ()
    spans: list[EvidenceSpan] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        spans.append(EvidenceSpan.model_validate_json(line))
    spans.sort(key=lambda item: item.id)
    return tuple(spans)


def _jsonl(models: Iterable[Any]) -> str:
    lines: list[str] = []
    for model in models:
        if hasattr(model, "model_dump"):
            lines.append(
                json.dumps(model.model_dump(mode="json"), ensure_ascii=False, sort_keys=True)
            )
        else:
            lines.append(json.dumps(model, ensure_ascii=False, sort_keys=True))
    return "\n".join(lines) + ("\n" if lines else "")


def write_routing_outputs(report: RoutingReport, output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "manifest": output_dir / "routing_manifest.json",
        "scenario_routes": output_dir / "scenario_routes.jsonl",
        "document_links": output_dir / "document_links.jsonl",
        "transaction_links": output_dir / "transaction_links.jsonl",
        "entity_conflicts": output_dir / "entity_conflicts.jsonl",
        "identity_evidence": output_dir / "identity_evidence.jsonl",
        "summary": output_dir / "routing_summary.md",
    }
    _atomic_write(paths["manifest"], report.manifest.model_dump_json(indent=2) + "\n")
    _atomic_write(paths["scenario_routes"], _jsonl(report.scenario_routes))
    _atomic_write(paths["document_links"], _jsonl(report.document_links))
    _atomic_write(paths["transaction_links"], _jsonl(report.transaction_links))
    _atomic_write(paths["entity_conflicts"], _jsonl(report.conflicts))
    _atomic_write(paths["identity_evidence"], _jsonl(report.identity_evidence))
    _atomic_write(paths["summary"], render_summary_markdown(report))
    return paths


def render_summary_markdown(report: RoutingReport) -> str:
    manifest = report.manifest
    account_dist = {route.scenario_id: len(route.account_ids) for route in report.scenario_routes}
    lines = [
        "# Routing summary (Stage 5B)",
        "",
        f"- scenarios: {manifest.scenario_count}",
        f"- template_cells: {manifest.template_cell_count}",
        f"- ledger_rows: {manifest.ledger_row_count}",
        f"- scenario_transactions: {manifest.scenario_transaction_count}",
        f"- transaction_links: {manifest.transaction_link_count}",
        f"- documents_resolved: {manifest.resolved_document_count}",
        f"- documents_unresolved: {manifest.unresolved_document_count}",
        f"- multi_scenario_documents: {manifest.multi_scenario_document_count}",
        f"- conflicts: {manifest.conflict_count}",
        f"- routing_algorithm_version: `{manifest.routing_algorithm_version}`",
        f"- normalization_version: `{manifest.normalization_version}`",
        "",
        "## Accounts per scenario",
        "",
    ]
    for route in report.scenario_routes:
        accounts = ", ".join(route.account_ids) if route.account_ids else "(none)"
        lines.append(
            f"- `{route.scenario_id}`: accounts=[{accounts}] "
            f"txns={route.transaction_count} docs={len(route.document_ids)}"
        )
    lines.extend(["", "## Account count distribution", ""])
    for scenario_id, count in sorted(account_dist.items()):
        lines.append(f"- `{scenario_id}`: {count}")
    lines.append("")
    return "\n".join(lines)


def has_structural_failure(report: RoutingReport) -> bool:
    """Strict-mode structural failures that make a scenario unusable."""
    for conflict in report.conflicts:
        if conflict.severity is DiagnosticSeverity.ERROR and conflict.kind in {
            ConflictKind.SCENARIO_WITHOUT_ACCOUNT,
            ConflictKind.DUPLICATE_TXN_ID,
            ConflictKind.TRANSACTION_ACCOUNT_CONFLICT,
        }:
            return True
    return False
