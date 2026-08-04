"""XLSX schema profiler using read-only openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from halyk_agent.adapters.archive.profilers.common import build_column_profile
from halyk_agent.config import Settings
from halyk_agent.domain.datasets import ArtifactFormat, SheetProfile, TableProfile


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def profile_xlsx(path: Path, *, artifact_id: str, settings: Settings) -> TableProfile:
    """Profile all sheets of an XLSX workbook without evaluating formulas."""
    warnings: list[str] = []
    size = path.stat().st_size
    if size > settings.max_profile_file_bytes:
        return TableProfile(
            artifact_id=artifact_id,
            format=ArtifactFormat.XLSX,
            sampled_rows=0,
            warnings=[
                f"XLSX larger than max_profile_file_bytes={settings.max_profile_file_bytes}; "
                "skipped deep profiling"
            ],
        )

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        return TableProfile(
            artifact_id=artifact_id,
            format=ArtifactFormat.XLSX,
            sampled_rows=0,
            warnings=[f"failed to open XLSX: {exc}"],
        )

    sheets: list[SheetProfile] = []
    total_sampled = 0
    try:
        for sheet in workbook.worksheets:
            rows_iter = sheet.iter_rows(values_only=True)
            sampled: list[tuple[Any, ...]] = []
            formula_count = 0
            for index, row in enumerate(rows_iter):
                # Count formulas from raw cells separately via another pass is costly in
                # read_only; approximate using string values starting with '='.
                formula_count += sum(1 for cell in row if _is_formula(cell))
                sampled.append(tuple(row))
                if index + 1 >= settings.max_sample_rows + 1:
                    break

            if not sampled:
                sheets.append(
                    SheetProfile(
                        name=sheet.title,
                        estimated_rows=0,
                        estimated_columns=0,
                        sampled_rows=0,
                        columns=[],
                        formula_cell_count=0,
                    )
                )
                warnings.append(f"empty sheet {sheet.title!r}")
                continue

            header_row = sampled[0]
            data_rows = sampled[1:] if len(sampled) > 1 else []
            header_detected = any(cell is not None and str(cell).strip() for cell in header_row)
            if header_detected:
                headers = [
                    str(cell).strip() if cell is not None and str(cell).strip() else f"column_{i}"
                    for i, cell in enumerate(header_row)
                ]
            else:
                headers = [f"column_{i}" for i in range(len(header_row))]
                data_rows = sampled
                warnings.append(f"header not detected on sheet {sheet.title!r}")

            width = len(headers)
            column_values: list[list[Any]] = [[] for _ in range(width)]
            for row in data_rows:
                for idx in range(width):
                    column_values[idx].append(row[idx] if idx < len(row) else None)

            columns = [
                build_column_profile(
                    name=headers[idx],
                    position=idx,
                    values=column_values[idx],
                    max_sample_value_length=settings.max_sample_value_length,
                )
                for idx in range(width)
            ]
            estimated_rows = sheet.max_row or len(sampled)
            estimated_cols = sheet.max_column or width
            sheets.append(
                SheetProfile(
                    name=sheet.title,
                    estimated_rows=int(estimated_rows),
                    estimated_columns=int(estimated_cols),
                    sampled_rows=len(data_rows),
                    columns=columns,
                    formula_cell_count=formula_count,
                )
            )
            total_sampled += len(data_rows)
    finally:
        workbook.close()

    return TableProfile(
        artifact_id=artifact_id,
        format=ArtifactFormat.XLSX,
        header_detected=True,
        sampled_rows=total_sampled,
        sheets=sheets,
        warnings=list(dict.fromkeys(warnings)),
    )
