"""Raw dataset discovery and quarantine (preflight process only)."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from collections.abc import Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from halyk_agent.config import Settings, get_settings
from halyk_agent.preflight.ignore import ignore_artifact
from halyk_agent.preflight.models import (
    AllowedInputRef,
    JsonCandidateRole,
    QuarantinedRef,
    SanitizedDatasetManifest,
)
from halyk_agent.preflight.quarantine import is_answer_key_payload
from halyk_agent.solver.errors import DatasetAdapterError

_LEDGER_ALIASES = {
    "txn_id": {"txn_id", "transaction_id", "transaction id", "txn id", "transactionid"},
    "date": {"date", "transaction_date", "transaction date", "txn_date", "posting_date"},
    "account_id": {"account_id", "account id", "account", "account_code", "account code"},
    "counterparty": {"counterparty", "counter_party", "vendor", "customer", "контрагент"},
    "description": {"description", "details", "narrative", "memo", "описание", "назначение"},
    "amount": {"amount", "value", "transaction_amount", "transaction amount", "сумма"},
    "currency": {"currency", "ccy", "currency_code", "currency code", "валюта"},
}


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _allowed(path: Path, data: bytes, role: str) -> AllowedInputRef:
    return AllowedInputRef(
        path=str(path.as_posix()),
        sha256=_sha256_bytes(data),
        size=len(data),
        role=role,
    )


def _quarantine(path: Path, data: bytes, reason: str) -> QuarantinedRef:
    return QuarantinedRef(
        path=str(path.as_posix()),
        sha256=_sha256_bytes(data),
        size=len(data),
        role=JsonCandidateRole.QUARANTINED_ANSWER_KEY,
        quarantine_reason=reason,
    )


def _clean_header(value: object) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[\s\-]+", "_", text)
    return text.strip("_")


def _canonical_header(header: Sequence[object]) -> set[str]:
    cleaned = {_clean_header(value) for value in header}
    canonical: set[str] = set()
    for target, aliases in _LEDGER_ALIASES.items():
        normalized_aliases = {_clean_header(alias) for alias in aliases}
        if cleaned & normalized_aliases:
            canonical.add(target)
    return canonical


def _csv_header(data: bytes) -> Sequence[object] | None:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        return None
    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    try:
        return next(csv.reader(io.StringIO(text), dialect=dialect))
    except StopIteration:
        return None


def _xlsx_header(data: bytes) -> Sequence[object] | None:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return None
    try:
        sheet = workbook.active
        if sheet is None:
            return None
        row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        return list(row) if row is not None else None
    finally:
        workbook.close()


def _looks_like_ledger(path: Path, data: bytes) -> bool:
    """Identify a ledger candidate from a minimal stable header fingerprint.

    Preflight discovery should identify candidates without duplicating the strict
    runtime ledger schema. The routing loader performs the full required-column
    validation before any candidate is used by the solver.
    """
    suffix = path.suffix.casefold()
    header: Sequence[object] | None
    if suffix in {".csv", ".txt", ".tsv"}:
        header = _csv_header(data)
    elif suffix in {".xlsx", ".xlsm"}:
        header = _xlsx_header(data)
    else:
        return False
    if header is None:
        return False
    required_fingerprint = {"txn_id", "amount", "currency"}
    return required_fingerprint.issubset(_canonical_header(header))


def _looks_like_submission_template(obj: Any) -> bool:
    if not isinstance(obj, dict):
        return False
    if not {"team", "contact_email", "model", "answers"}.issubset(obj):
        return False
    answers = obj.get("answers")
    if not isinstance(answers, dict) or not answers:
        return False
    first = next(iter(answers.values()))
    if not isinstance(first, dict) or not first:
        return False
    cell = next(iter(first.values()))
    return isinstance(cell, dict) and {"status", "actual", "evidence_txn_id"}.issubset(cell)


def _looks_like_case_markdown(path: Path, data: bytes) -> bool:
    if path.suffix.casefold() not in {".md", ".markdown", ".txt"}:
        return False
    try:
        text = data.decode("utf-8-sig").casefold()
    except UnicodeDecodeError:
        return False
    signals = ("covenant", "ковенант", "scenario", "сценари", "лимит", "limit")
    return sum(1 for signal in signals if signal in text) >= 2


def _check_input_limits(root: Path, files: list[Path], settings: Settings) -> None:
    if len(files) > settings.max_archive_files:
        raise DatasetAdapterError(
            f"dataset contains {len(files)} files, exceeding limit {settings.max_archive_files}"
        )
    total = 0
    for path in files:
        size = path.stat().st_size
        if size > settings.max_single_file_bytes:
            raise DatasetAdapterError(
                f"dataset file exceeds size limit: {path.name} ({size} > {settings.max_single_file_bytes})"
            )
        total += size
        if total > settings.max_total_uncompressed_bytes:
            raise DatasetAdapterError(
                "dataset total bytes exceed configured uncompressed-size limit"
            )


def discover_and_sanitize(
    root: Path,
    *,
    settings: Settings | None = None,
) -> tuple[SanitizedDatasetManifest, list[dict[str, str]]]:
    """Inspect raw dataset and return the minimal solver allowlist."""
    root = root.resolve()
    if not root.is_dir():
        raise DatasetAdapterError(f"dataset root is not a directory: {root}")
    resolved_settings = settings or get_settings()

    inspected: list[dict[str, str]] = []
    ignored = []
    case_descriptions: list[AllowedInputRef] = []
    technical_noise: list[AllowedInputRef] = []
    document_files: list[AllowedInputRef] = []
    ledgers: list[AllowedInputRef] = []
    templates: list[AllowedInputRef] = []
    quarantined: list[QuarantinedRef] = []
    documents_dir: Path | None = None

    all_files = sorted(
        (path for path in root.rglob("*") if path.is_file()), key=lambda p: p.as_posix()
    )
    _check_input_limits(root, all_files, resolved_settings)

    for path in all_files:
        ignored_item = ignore_artifact(path)
        if ignored_item is not None:
            ignored.append(ignored_item)
            inspected.append(
                {
                    "path": ignored_item.path,
                    "role": "IGNORED",
                    "sha256": ignored_item.sha256,
                    "size": str(ignored_item.size),
                    "note": ignored_item.ignore_rule,
                }
            )
            continue

        data = path.read_bytes()
        rel_parent = path.parent

        if path.suffix.casefold() == ".json":
            name_l = path.name.casefold()
            if "ground_truth" in name_l:
                item = _quarantine(path, data, "filename_ground_truth")
                quarantined.append(item)
                inspected.append(
                    {
                        "path": item.path,
                        "role": item.role.value,
                        "sha256": item.sha256,
                        "size": str(item.size),
                        "note": item.quarantine_reason,
                    }
                )
                continue
            try:
                obj = json.loads(data.decode("utf-8-sig"))
            except (UnicodeDecodeError, json.JSONDecodeError):
                obj = None
            if obj is not None and is_answer_key_payload(obj):
                item = _quarantine(path, data, "content_shape_answer_key")
                quarantined.append(item)
                inspected.append(
                    {
                        "path": item.path,
                        "role": item.role.value,
                        "sha256": item.sha256,
                        "size": str(item.size),
                        "note": item.quarantine_reason,
                    }
                )
                continue
            if obj is not None and _looks_like_submission_template(obj):
                ref = _allowed(path, data, "submission_template")
                templates.append(ref)
                inspected.append(
                    {
                        "path": ref.path,
                        "role": JsonCandidateRole.SUBMISSION_TEMPLATE.value,
                        "sha256": ref.sha256,
                        "size": str(ref.size),
                        "note": "allowlisted",
                    }
                )
                continue
            inspected.append(
                {
                    "path": str(path.as_posix()),
                    "role": JsonCandidateRole.UNKNOWN_JSON.value,
                    "sha256": _sha256_bytes(data),
                    "size": str(len(data)),
                    "note": "not_allowlisted",
                }
            )
            continue

        if _looks_like_ledger(path, data):
            ledgers.append(_allowed(path, data, "primary_ledger"))
            continue

        if _looks_like_case_markdown(path, data):
            case_descriptions.append(_allowed(path, data, "case_description"))
            continue

        if path.suffix.casefold() == ".pdf":
            document_files.append(_allowed(path, data, "document"))
            if documents_dir is None or len(list(rel_parent.glob("*.pdf"))) > len(
                list((documents_dir or rel_parent).glob("*.pdf"))
            ):
                documents_dir = rel_parent
            continue

        if "documents" in {part.casefold() for part in path.parts}:
            technical_noise.append(_allowed(path, data, "technical_noise"))

    if not templates:
        raise DatasetAdapterError("submission template not found")
    if not ledgers:
        raise DatasetAdapterError("primary ledger not found")
    if not case_descriptions:
        raise DatasetAdapterError("case description not found")
    if not document_files:
        raise DatasetAdapterError("document files not found")

    root_templates = [item for item in templates if Path(item.path).parent.resolve() == root]
    if len(root_templates) > 1:
        raise DatasetAdapterError("ambiguous submission templates at dataset root")
    if root_templates:
        submission_template = root_templates[0]
    elif len(templates) == 1:
        submission_template = templates[0]
    else:
        raise DatasetAdapterError("ambiguous submission templates")

    if len(ledgers) != 1:
        raise DatasetAdapterError(f"expected exactly one primary ledger, found {len(ledgers)}")

    if documents_dir is None:
        documents_dir = root

    manifest = SanitizedDatasetManifest(
        case_descriptions=case_descriptions,
        primary_ledger=ledgers[0],
        submission_template=submission_template,
        documents_dir=documents_dir.resolve().as_posix(),
        document_files=document_files,
        technical_noise=technical_noise,
        ignored=ignored,
        quarantined=quarantined,
    )
    return manifest, inspected
